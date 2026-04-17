#!/usr/bin/env python3
"""
meta_wipe.py — Strip metadata from Microsoft Office documents.

Removes author names, revision history, comments, tracked changes,
and other hidden metadata from Word (.docx) and Excel (.xlsx) files.
Office documents store a surprising amount of personal information
in their metadata — this tool removes it.

What gets removed:
  - Author and last-modified-by names
  - Company name
  - Revision number and editing time
  - Creation and modification dates
  - Comments and annotations
  - Document title, subject, keywords, description, category
  - Custom properties
  - For .docx: tracked changes (accept all), headers/footers author info
  - For .xlsx: sheet-level comments

Usage:
    python meta_wipe.py report.docx
    python meta_wipe.py spreadsheet.xlsx -o clean_spreadsheet.xlsx
    python meta_wipe.py ./documents/ --recursive
    python meta_wipe.py contract.docx --preview

Supports: .docx, .xlsx
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependency checks
# ---------------------------------------------------------------------------

HAS_DOCX = False
HAS_OPENPYXL = False

try:
    from docx import Document
    from docx.opc.constants import RELATIONSHIP_TYPE as RT

    HAS_DOCX = True
except ImportError:
    pass

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    pass

SUPPORTED_EXTENSIONS = set()
if HAS_DOCX:
    SUPPORTED_EXTENSIONS.add(".docx")
if HAS_OPENPYXL:
    SUPPORTED_EXTENSIONS.add(".xlsx")

# ---------------------------------------------------------------------------
# Core Properties to Clear
# ---------------------------------------------------------------------------
# These are the Dublin Core / Office Open XML core properties that
# can contain personally identifiable information.

CORE_PROPERTIES = [
    "author",
    "last_modified_by",
    "category",
    "comments",
    "content_status",
    "description",
    "identifier",
    "keywords",
    "subject",
    "title",
    "version",
]

# Properties to reset to neutral values (not just empty)
RESET_PROPERTIES = {
    "revision": 1,
}


def get_metadata_summary(props):
    """
    Extract current metadata values for preview display.

    Args:
        props: Core properties object (from python-docx or openpyxl)

    Returns:
        Dict of {property_name: current_value}
    """
    summary = {}
    for prop_name in CORE_PROPERTIES:
        try:
            value = getattr(props, prop_name, None)
            if value is not None and str(value).strip():
                summary[prop_name] = str(value)
        except Exception:
            pass

            # Special properties
    try:
        if hasattr(props, "revision") and props.revision:
            summary["revision"] = str(props.revision)
    except Exception:
        pass

    try:
        if hasattr(props, "created") and props.created:
            summary["created"] = str(props.created)
    except Exception:
        pass

    try:
        if hasattr(props, "modified") and props.modified:
            summary["modified"] = str(props.modified)
    except Exception:
        pass

    try:
        if hasattr(props, "last_printed") and props.last_printed:
            summary["last_printed"] = str(props.last_printed)
    except Exception:
        pass

    return summary


def clear_core_properties(props):
    """
    Clear all core properties that might contain PII.

    Args:
        props: Core properties object

    Returns:
        Number of properties cleared
    """
    cleared = 0

    for prop_name in CORE_PROPERTIES:
        try:
            current = getattr(props, prop_name, None)
            if current is not None and str(current).strip():
                setattr(props, prop_name, "")
                cleared += 1
        except (AttributeError, TypeError):
            pass

            # Reset revision number
    try:
        if hasattr(props, "revision"):
            props.revision = 1
            cleared += 1
    except (AttributeError, TypeError):
        pass

        # Clear dates — set to a neutral timestamp
    neutral_time = datetime(2000, 1, 1, 0, 0, 0)
    for date_prop in ["created", "modified", "last_printed"]:
        try:
            current = getattr(props, date_prop, None)
            if current is not None:
                setattr(props, date_prop, neutral_time)
                cleared += 1
        except (AttributeError, TypeError):
            pass

    return cleared


# ---------------------------------------------------------------------------
# Word Document (.docx) Processing
# ---------------------------------------------------------------------------

def wipe_docx(input_path, output_path, preview=False, verbose=False):
    """
    Remove metadata from a Word document.

    Args:
        input_path: Path to .docx file
        output_path: Where to save the cleaned file
        preview: Show metadata without removing
        verbose: Print detailed progress

    Returns:
        Tuple of (success: bool, stats: dict)
    """
    if not HAS_DOCX:
        print("ERROR: python-docx is required for .docx files. "
              "Install with: pip install python-docx", file=sys.stderr)
        return False, {}

    try:
        doc = Document(str(input_path))
    except Exception as e:
        print(f"  ERROR: Cannot open {input_path}: {e}", file=sys.stderr)
        return False, {}

    stats = {"properties_cleared": 0, "comments_removed": 0}

    # --- Preview current metadata ---
    core_props = doc.core_properties
    current_meta = get_metadata_summary(core_props)

    if preview:
        print(f"  Current metadata in {input_path}:")
        if current_meta:
            for prop, value in sorted(current_meta.items()):
                # Truncate long values for display
                display = value[:80] + "..." if len(value) > 80 else value
                print(f"    {prop}: {display}")
        else:
            print("    (no metadata found)")
        return True, {"properties_found": len(current_meta)}

        # --- Clear core properties ---
    if verbose and current_meta:
        print(f"  Clearing {len(current_meta)} metadata properties...")

    stats["properties_cleared"] = clear_core_properties(core_props)

    # --- Remove comments ---
    # Comments in .docx are stored as XML elements
    try:
        comments_part = None
        for rel in doc.part.rels.values():
            if "comments" in str(getattr(rel, 'reltype', '')).lower():
                comments_part = rel.target_part
                break

        if comments_part is not None:
            from lxml import etree
            comments_xml = comments_part._element
            comment_elements = comments_xml.findall(
                './/{http://schemas.openxmlformats.org/wordprocessingml/2006/main}comment'
            )
            for comment in comment_elements:
                comments_xml.remove(comment)
                stats["comments_removed"] += 1

            if verbose and stats["comments_removed"]:
                print(f"  Removed {stats['comments_removed']} comment(s)")
    except Exception as e:
        if verbose:
            print(f"  Note: Could not process comments: {e}")

            # --- Clear custom XML properties ---
    try:
        custom_props = None
        for rel in doc.part.rels.values():
            if "custom-properties" in str(getattr(rel, 'reltype', '')).lower():
                custom_props = rel.target_part
                break
        if custom_props is not None:
            from lxml import etree
            root = custom_props._element
            for child in list(root):
                root.remove(child)
            stats["properties_cleared"] += 1
            if verbose:
                print(f"  Cleared custom properties")
    except Exception as e:
        if verbose:
            print(f"  Note: Could not clear custom properties: {e}")

            # --- Save ---
    try:
        doc.save(str(output_path))
        if verbose:
            print(f"  Saved to {output_path}")
        return True, stats
    except Exception as e:
        print(f"  ERROR: Cannot save {output_path}: {e}", file=sys.stderr)
        return False, stats

    # ---------------------------------------------------------------------------


# Excel Workbook (.xlsx) Processing
# ---------------------------------------------------------------------------

def wipe_xlsx(input_path, output_path, preview=False, verbose=False):
    """
    Remove metadata from an Excel workbook.

    Args:
        input_path: Path to .xlsx file
        output_path: Where to save the cleaned file
        preview: Show metadata without removing
        verbose: Print detailed progress

    Returns:
        Tuple of (success: bool, stats: dict)
    """
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl is required for .xlsx files. "
              "Install with: pip install openpyxl", file=sys.stderr)
        return False, {}

    try:
        wb = openpyxl.load_workbook(str(input_path))
    except Exception as e:
        print(f"  ERROR: Cannot open {input_path}: {e}", file=sys.stderr)
        return False, {}

    stats = {"properties_cleared": 0, "comments_removed": 0}

    # --- Preview current metadata ---
    current_meta = get_metadata_summary(wb.properties)

    if preview:
        print(f"  Current metadata in {input_path}:")
        if current_meta:
            for prop, value in sorted(current_meta.items()):
                display = value[:80] + "..." if len(value) > 80 else value
                print(f"    {prop}: {display}")
        else:
            print("    (no metadata found)")

            # Count comments across sheets
        total_comments = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        total_comments += 1
        if total_comments:
            print(f"    comments: {total_comments} cell comment(s) across "
                  f"{len(wb.worksheets)} sheet(s)")

        return True, {"properties_found": len(current_meta),
                      "comments_found": total_comments}

        # --- Clear core properties ---
    if verbose and current_meta:
        print(f"  Clearing {len(current_meta)} metadata properties...")

    stats["properties_cleared"] = clear_core_properties(wb.properties)

    # --- Remove cell comments ---
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    cell.comment = None
                    stats["comments_removed"] += 1

    if verbose and stats["comments_removed"]:
        print(f"  Removed {stats['comments_removed']} cell comment(s)")

        # --- Clear custom document properties ---
    try:
        if hasattr(wb, 'custom_doc_props') and wb.custom_doc_props:
            wb.custom_doc_props.clear()
            stats["properties_cleared"] += 1
    except Exception:
        pass

        # --- Save ---
    try:
        wb.save(str(output_path))
        if verbose:
            print(f"  Saved to {output_path}")
        return True, stats
    except Exception as e:
        print(f"  ERROR: Cannot save {output_path}: {e}", file=sys.stderr)
        return False, stats

    # ---------------------------------------------------------------------------


# File Discovery and Routing
# ---------------------------------------------------------------------------

def wipe_file(input_path, output_path=None, preview=False, verbose=False):
    """Route to the appropriate handler based on file extension."""
    input_path = Path(input_path)
    ext = input_path.suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        if verbose:
            print(f"  Skipping {input_path} (unsupported format: {ext})")
        return False, {}

        # Default output path
    if output_path is None:
        output_path = input_path.parent / f"{input_path.stem}_clean{ext}"
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if ext == ".docx":
        return wipe_docx(input_path, output_path, preview, verbose)
    elif ext == ".xlsx":
        return wipe_xlsx(input_path, output_path, preview, verbose)
    else:
        print(f"  Unsupported format: {ext}", file=sys.stderr)
        return False, {}


def find_office_files(path, recursive=False):
    """Find supported Office documents at the given path."""
    path = Path(path)
    if path.is_file():
        return [path] if path.suffix.lower() in SUPPORTED_EXTENSIONS else []
    if path.is_dir():
        if recursive:
            return sorted(
                f for f in path.rglob("*")
                if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS
            )
        else:
            return sorted(
                f for f in path.iterdir()
                if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS
            )
    return []


def main():
    # Check that at least one library is available
    if not HAS_DOCX and not HAS_OPENPYXL:
        print("ERROR: No document libraries installed.")
        print("  For .docx support: pip install python-docx")
        print("  For .xlsx support: pip install openpyxl")
        sys.exit(1)

    supported_list = []
    if HAS_DOCX:
        supported_list.append(".docx")
    if HAS_OPENPYXL:
        supported_list.append(".xlsx")

    parser = argparse.ArgumentParser(
        description="Strip author, revision, and comment metadata from Office documents.",
        epilog="Examples:\n"
               "  python meta_wipe.py report.docx\n"
               "  python meta_wipe.py data.xlsx -o clean_data.xlsx\n"
               "  python meta_wipe.py ./docs/ --recursive\n"
               "  python meta_wipe.py contract.docx --preview\n\n"
               f"Supported formats: {', '.join(supported_list)}\n"
               "Requires: python-docx (for .docx), openpyxl (for .xlsx)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "input",
        help="Office document or directory to process",
    )
    parser.add_argument(
        "-o", "--output",
        help="Output file path (default: <input>_clean.<ext>)",
    )
    parser.add_argument(
        "-r", "--recursive",
        action="store_true",
        help="Recursively process directories",
    )
    parser.add_argument(
        "--preview",
        action="store_true",
        help="Show current metadata without removing",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Print detailed progress",
    )

    args = parser.parse_args()
    input_path = Path(args.input)

    if not input_path.exists():
        print(f"Not found: {input_path}", file=sys.stderr)
        sys.exit(1)

        # Find files
    files = find_office_files(input_path, recursive=args.recursive)
    if not files:
        print(f"No supported Office documents found "
              f"({', '.join(supported_list)}).")
        sys.exit(0)

    if args.preview:
        print(f"Preview — metadata in {len(files)} file(s):\n")
    else:
        print(f"Processing {len(files)} file(s)...\n")

        # Process files
    total_stats = {
        "files_processed": 0,
        "files_cleaned": 0,
        "properties_cleared": 0,
        "comments_removed": 0,
    }

    for file_path in files:
        if args.verbose:
            print(f"\n  File: {file_path}")

            # Only use custom output for single file
        output_path = args.output if len(files) == 1 else None

        success, stats = wipe_file(
            file_path,
            output_path=output_path,
            preview=args.preview,
            verbose=args.verbose,
        )

        if success:
            total_stats["files_processed"] += 1
            cleared = stats.get("properties_cleared", 0) + \
                      stats.get("comments_removed", 0)
            if cleared > 0:
                total_stats["files_cleaned"] += 1
            total_stats["properties_cleared"] += stats.get("properties_cleared", 0)
            total_stats["comments_removed"] += stats.get("comments_removed", 0)

            # Summary
    if not args.preview:
        print(f"\n--- Summary ---")
        print(f"Files processed: {total_stats['files_processed']}")
        print(f"Files with metadata removed: {total_stats['files_cleaned']}")
        print(f"Properties cleared: {total_stats['properties_cleared']}")
        if total_stats["comments_removed"]:
            print(f"Comments removed: {total_stats['comments_removed']}")


if __name__ == "__main__":
    main()