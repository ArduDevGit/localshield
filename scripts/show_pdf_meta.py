#!/usr/bin/env python3
"""
meta_scan.py — LocalShield Metadata Scanner

Scans files for embedded metadata (author names, software, timestamps,
GPS coordinates, etc.) WITHOUT reading file content. Supports PDFs,
images (JPEG/PNG/TIFF/WebP), Word documents (.docx), Excel spreadsheets
(.xlsx), and reports OS-level metadata for all file types.

Part of the LocalShield Privacy Toolkit.
"""

import argparse
import json
import os
import stat
import sys
import time
from datetime import datetime
from pathlib import Path

# Optional imports — degrade gracefully if a library is missing
try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    from PIL import Image
    from PIL.ExifTags import TAGS, GPSTAGS
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ──────────────────────────────────────────────
# Color helpers (disabled when piped / Windows)
# ──────────────────────────────────────────────
USE_COLOR = sys.stdout.isatty() and os.name != "nt"

def _c(code, text):
    return f"\033[{code}m{text}\033[0m" if USE_COLOR else text

def red(t):    return _c("31", t)
def yellow(t): return _c("33", t)
def green(t):  return _c("32", t)
def cyan(t):   return _c("36", t)
def bold(t):   return _c("1", t)
def dim(t):    return _c("2", t)


# ──────────────────────────────────────────────
# Risk assessment — flag fields that may leak PII
# ──────────────────────────────────────────────
PII_FIELDS = {
    "author", "creator", "producer", "last_modified_by", "company",
    "manager", "gps_latitude", "gps_longitude", "gps_position",
    "artist", "copyright", "camera_owner", "body_serial_number",
    "lens_serial_number", "software", "host_computer",
}

def risk_level(key, value):
    """Return 'high', 'medium', or 'low' risk for a metadata field."""
    k = key.lower().replace(" ", "_")
    if value is None or str(value).strip() == "":
        return "low"
    if k in PII_FIELDS:
        return "high"
    if any(word in k for word in ("date", "time", "modified", "created")):
        return "medium"
    return "low"

def risk_icon(level):
    if level == "high":
        return red("▲ HIGH")
    elif level == "medium":
        return yellow("● MED ")
    return green("○ LOW ")


# ──────────────────────────────────────────────
# OS-level metadata (works for every file)
# ──────────────────────────────────────────────
def scan_os_metadata(filepath):
    """Return OS-level file metadata."""
    meta = {}
    try:
        st = os.stat(filepath)
        meta["File Size"] = _format_size(st.st_size)
        meta["Last Modified"] = _fmt_time(st.st_mtime)
        meta["Last Accessed"] = _fmt_time(st.st_atime)
        if hasattr(st, "st_birthtime"):
            meta["Created"] = _fmt_time(st.st_birthtime)
        elif hasattr(st, "st_ctime"):
            meta["Created (ctime)"] = _fmt_time(st.st_ctime)
        meta["Permissions"] = stat.filemode(st.st_mode)
    except OSError as e:
        meta["Error"] = str(e)
    return meta


# ──────────────────────────────────────────────
# PDF metadata
# ──────────────────────────────────────────────
def scan_pdf(filepath):
    """Extract metadata from a PDF file using PyMuPDF."""
    if not HAS_FITZ:
        return {"Warning": "PyMuPDF (fitz) not installed — pip install PyMuPDF"}
    meta = {}
    try:
        doc = fitz.open(filepath)
        raw = doc.metadata or {}
        field_map = {
            "author": "Author",
            "creator": "Creator",
            "producer": "Producer",
            "title": "Title",
            "subject": "Subject",
            "keywords": "Keywords",
            "creationDate": "Creation Date",
            "modDate": "Modification Date",
            "format": "PDF Format",
            "encryption": "Encryption",
        }
        for key, label in field_map.items():
            val = raw.get(key, "")
            if val:
                meta[label] = val
        meta["Page Count"] = doc.page_count
        doc.close()
    except Exception as e:
        meta["Error"] = str(e)
    return meta


# ──────────────────────────────────────────────
# Image metadata (EXIF + GPS)
# ──────────────────────────────────────────────
def scan_image(filepath):
    """Extract EXIF and GPS metadata from an image file."""
    if not HAS_PILLOW:
        return {"Warning": "Pillow not installed — pip install Pillow"}
    meta = {}
    try:
        img = Image.open(filepath)
        meta["Format"] = img.format
        meta["Dimensions"] = f"{img.width} x {img.height}"
        meta["Color Mode"] = img.mode

        exif_data = img.getexif()
        if not exif_data:
            meta["EXIF"] = "None found"
            return meta

        for tag_id, value in exif_data.items():
            tag_name = TAGS.get(tag_id, f"Unknown-{tag_id}")
            # Skip thumbnail and binary blobs
            if isinstance(value, bytes) and len(value) > 100:
                meta[tag_name] = f"<binary data, {len(value)} bytes>"
            else:
                meta[tag_name] = str(value)

        # GPS data (nested IFD)
        gps_ifd = exif_data.get_ifd(0x8825)
        if gps_ifd:
            gps_meta = {}
            for gps_tag_id, gps_val in gps_ifd.items():
                gps_tag_name = GPSTAGS.get(gps_tag_id, f"Unknown-{gps_tag_id}")
                gps_meta[gps_tag_name] = str(gps_val)
            lat = _parse_gps_coord(gps_ifd, "GPSLatitude", "GPSLatitudeRef")
            lon = _parse_gps_coord(gps_ifd, "GPSLongitude", "GPSLongitudeRef")
            if lat is not None and lon is not None:
                meta["GPS Latitude"] = f"{lat:.6f}"
                meta["GPS Longitude"] = f"{lon:.6f}"
                meta["GPS Position"] = f"{lat:.6f}, {lon:.6f}"
            for k, v in gps_meta.items():
                if k not in ("GPSLatitude", "GPSLongitude", "GPSLatitudeRef", "GPSLongitudeRef"):
                    meta[f"GPS {k}"] = v

        img.close()
    except Exception as e:
        meta["Error"] = str(e)
    return meta


def _parse_gps_coord(gps_ifd, coord_key, ref_key):
    """Convert GPS coordinate tuple to decimal degrees."""
    try:
        coord_tag = {v: k for k, v in GPSTAGS.items()}
        coord = gps_ifd.get(coord_tag.get(coord_key))
        ref = gps_ifd.get(coord_tag.get(ref_key))
        if coord and ref:
            degrees = float(coord[0])
            minutes = float(coord[1])
            seconds = float(coord[2])
            decimal = degrees + minutes / 60.0 + seconds / 3600.0
            if ref in ("S", "W"):
                decimal = -decimal
            return decimal
    except (TypeError, IndexError, ValueError, ZeroDivisionError):
        pass
    return None


# ──────────────────────────────────────────────
# Word (.docx) metadata
# ──────────────────────────────────────────────
def scan_docx(filepath):
    """Extract metadata from a Word document."""
    if not HAS_DOCX:
        return {"Warning": "python-docx not installed — pip install python-docx"}
    meta = {}
    try:
        doc = DocxDocument(filepath)
        props = doc.core_properties
        field_map = [
            ("Author", props.author),
            ("Last Modified By", props.last_modified_by),
            ("Title", props.title),
            ("Subject", props.subject),
            ("Keywords", props.keywords),
            ("Category", props.category),
            ("Comments", props.comments),
            ("Revision", props.revision),
            ("Created", props.created),
            ("Modified", props.modified),
            ("Last Printed", props.last_printed),
            ("Content Status", props.content_status),
            ("Language", props.language),
            ("Identifier", props.identifier),
            ("Version", props.version),
        ]
        for label, value in field_map:
            if value is not None and str(value).strip():
                meta[label] = str(value)
    except Exception as e:
        meta["Error"] = str(e)
    return meta


# ──────────────────────────────────────────────
# Excel (.xlsx) metadata
# ──────────────────────────────────────────────
def scan_xlsx(filepath):
    """Extract metadata from an Excel workbook."""
    if not HAS_OPENPYXL:
        return {"Warning": "openpyxl not installed — pip install openpyxl"}
    meta = {}
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        props = wb.properties
        field_map = [
            ("Creator", props.creator),
            ("Last Modified By", props.lastModifiedBy),
            ("Title", props.title),
            ("Subject", props.subject),
            ("Keywords", props.keywords),
            ("Category", props.category),
            ("Description", props.description),
            ("Created", props.created),
            ("Modified", props.modified),
            ("Last Printed", props.lastPrinted),
            ("Company", props.company if hasattr(props, "company") else None),
            ("Manager", props.manager if hasattr(props, "manager") else None),
            ("Version", props.version),
        ]
        for label, value in field_map:
            if value is not None and str(value).strip():
                meta[label] = str(value)
        meta["Sheet Count"] = len(wb.sheetnames)
        meta["Sheet Names"] = ", ".join(wb.sheetnames)
        wb.close()
    except Exception as e:
        meta["Error"] = str(e)
    return meta


# ──────────────────────────────────────────────
# File type routing
# ──────────────────────────────────────────────
EXT_MAP = {
    ".pdf": ("PDF Document", scan_pdf),
    ".jpg": ("Image (JPEG)", scan_image),
    ".jpeg": ("Image (JPEG)", scan_image),
    ".png": ("Image (PNG)", scan_image),
    ".tiff": ("Image (TIFF)", scan_image),
    ".tif": ("Image (TIFF)", scan_image),
    ".webp": ("Image (WebP)", scan_image),
    ".bmp": ("Image (BMP)", scan_image),
    ".gif": ("Image (GIF)", scan_image),
    ".docx": ("Word Document", scan_docx),
    ".xlsx": ("Excel Spreadsheet", scan_xlsx),
}

def scan_file(filepath):
    """Scan a single file and return (file_type, os_meta, format_meta)."""
    ext = Path(filepath).suffix.lower()
    os_meta = scan_os_metadata(filepath)

    if ext in EXT_MAP:
        file_type, scanner = EXT_MAP[ext]
        format_meta = scanner(filepath)
    else:
        file_type = f"Other ({ext or 'no extension'})"
        format_meta = {}

    return file_type, os_meta, format_meta


# ──────────────────────────────────────────────
# Output formatting
# ──────────────────────────────────────────────
def print_report(filepath, file_type, os_meta, format_meta, verbose=False):
    """Print a formatted metadata report for one file."""
    print()
    print(bold(f"{'═' * 60}"))
    print(bold(f"  {filepath}"))
    print(dim(f"  Type: {file_type}"))
    print(bold(f"{'═' * 60}"))

    # OS metadata
    if verbose:
        print(cyan("\n  ── OS Metadata ──"))
        for k, v in os_meta.items():
            lvl = risk_level(k, v)
            print(f"  {risk_icon(lvl)}  {k}: {v}")

    # Format-specific metadata
    if format_meta:
        print(cyan(f"\n  ── {file_type} Metadata ──"))
        has_high = False
        for k, v in format_meta.items():
            lvl = risk_level(k, v)
            if lvl == "high":
                has_high = True
            print(f"  {risk_icon(lvl)}  {k}: {v}")
        if has_high:
            print()
            print(red("  ⚠  HIGH-RISK metadata found — may contain personal information"))
    elif not format_meta and file_type.startswith("Other"):
        print(dim("\n  No format-specific metadata scanner for this file type."))
        print(dim("  OS-level metadata shown with --verbose."))
    else:
        print(green("\n  ✓ No embedded metadata found"))

    print()


def print_summary(results):
    """Print a summary table of all scanned files."""
    print(bold(f"\n{'═' * 60}"))
    print(bold("  SCAN SUMMARY"))
    print(bold(f"{'═' * 60}"))

    high_count = 0
    med_count = 0
    clean_count = 0

    for filepath, (file_type, os_meta, format_meta) in results.items():
        max_risk = "low"
        for k, v in format_meta.items():
            lvl = risk_level(k, v)
            if lvl == "high":
                max_risk = "high"
                break
            elif lvl == "medium":
                max_risk = "medium"

        if max_risk == "high":
            high_count += 1
            icon = red("▲ HIGH")
        elif max_risk == "medium":
            med_count += 1
            icon = yellow("● MED ")
        else:
            clean_count += 1
            icon = green("○ CLEAN")

        print(f"  {icon}  {filepath}")

    print(f"\n  Files scanned: {len(results)}")
    if high_count:
        print(red(f"  High risk:     {high_count}"))
    if med_count:
        print(yellow(f"  Medium risk:   {med_count}"))
    print(green(f"  Clean:         {clean_count}"))
    print()


def export_json(results, output_path):
    """Export scan results as JSON."""
    export = {}
    for filepath, (file_type, os_meta, format_meta) in results.items():
        export[filepath] = {
            "file_type": file_type,
            "os_metadata": os_meta,
            "format_metadata": format_meta,
            "fields": {
                k: {"value": v, "risk": risk_level(k, v)}
                for k, v in {**os_meta, **format_meta}.items()
            }
        }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(export, f, indent=2, default=str)
    print(green(f"\n  ✓ Results exported to {output_path}"))


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
def _format_size(size_bytes):
    for unit in ("B", "KB", "MB", "GB"):
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"

def _fmt_time(timestamp):
    try:
        return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
    except (OSError, ValueError):
        return str(timestamp)


def collect_files(targets, recursive=False):
    """Resolve targets (files and/or directories) into a list of file paths."""
    files = []
    for target in targets:
        p = Path(target)
        if p.is_file():
            files.append(str(p))
        elif p.is_dir():
            if recursive:
                for child in sorted(p.rglob("*")):
                    if child.is_file():
                        files.append(str(child))
            else:
                for child in sorted(p.iterdir()):
                    if child.is_file():
                        files.append(str(child))
        else:
            print(yellow(f"  ⚠ Skipping (not found): {target}"))
    return files


# ──────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        prog="meta_scan.py",
        description="Scan files for embedded metadata that may leak personal information.",
        epilog="Part of the LocalShield Privacy Toolkit — https://github.com/ArduDevGit/localshield",
    )
    parser.add_argument(
        "targets",
        nargs="+",
        help="Files or directories to scan",
    )
    parser.add_argument(
        "-r", "--recursive",
        action="store_true",
        help="Recursively scan directories",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Show OS-level metadata for every file",
    )
    parser.add_argument(
        "-o", "--output",
        metavar="FILE",
        help="Export results as JSON to FILE",
    )
    parser.add_argument(
        "--summary-only",
        action="store_true",
        help="Show only the summary table, not per-file details",
    )

    args = parser.parse_args()

    files = collect_files(args.targets, args.recursive)

    if not files:
        print(red("\n  ✗ No files found to scan.\n"))
        sys.exit(1)

    print(bold(f"\n  Scanning {len(files)} file(s) for metadata...\n"))

    results = {}
    for filepath in files:
        file_type, os_meta, format_meta = scan_file(filepath)
        results[filepath] = (file_type, os_meta, format_meta)
        if not args.summary_only:
            print_report(filepath, file_type, os_meta, format_meta, args.verbose)

    print_summary(results)

    if args.output:
        export_json(results, args.output)


if __name__ == "__main__":
    main()
